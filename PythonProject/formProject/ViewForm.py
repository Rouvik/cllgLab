import ttkbootstrap as ttk
from tkinter import messagebox, filedialog
import openpyxl

class ViewForm(ttk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        self.glb_row = 1

        ttk.Label(master=self, text="Pick student data file: ").grid(row=0, column=0)
        ttk.Button(master=self, text="Pick File", command=self.showStudentData).grid(row=0, column=1, pady=30)

        self.name = self.genInfoLine("Name: ")
        self.addr = self.genInfoLine("Address: ")
        self.dob = self.genInfoLine("Date of Birth: ")
        self.dept = self.genInfoLine("Department: ")
        self.cls10 = self.genInfoLine("Class 10 marks (%): ")
        self.cls12 = self.genInfoLine("Class 12 marks (%): ")
        self.gender = self.genInfoLine("Gender: ")

        ttk.Button(master=self, text="Back", command=lambda: controller.showFrame("MainForm")).grid(row=self.glb_row, column=1)

    def genInfoLine(self, label):
        ttk.Label(master=self, text=label, font=("Helvetica", 15)).grid(row=self.glb_row, pady=10)
        x = ttk.Label(master=self, font=("Helvetica", 15))
        x.grid(row=self.glb_row, column=1)
        self.glb_row += 1
        return x
    
    def updateData(self, ws):
        self.name.configure(text=ws[2][0].value)
        self.addr.configure(text=ws[2][1].value)
        self.dob.configure(text=ws[2][2].value)
        self.dept.configure(text=ws[2][3].value)
        self.cls10.configure(text=ws[2][4].value)
        self.cls12.configure(text=ws[2][5].value)
        self.gender.configure(text=ws[2][6].value)

    def showStudentData(self):
        fname = filedialog.askopenfilename()
        if fname is None or len(fname) == 0:
            messagebox.showerror(message="Error failed to open file to read")
            return
        
        wb = openpyxl.load_workbook(fname)
        ws = wb.active
        self.updateData(ws)